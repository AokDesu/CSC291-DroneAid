"""Build docs/02-backlog-delphi.xlsx — Wideband Delphi mock-filled backlog."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from pathlib import Path

# Fibonacci scale used for estimates
FIB = [1, 2, 3, 5, 8, 13]

# (id, epic, story, acceptance, owner, priority, sprint, r1, r2, r3)
# r1/r2/r3 are lists of 5 estimator values from [Aok, Belle, Bew, Poom, Tawan]
# Mock data shows realistic Delphi convergence: spread shrinks each round.
STORIES = [
    # ── Identity ──
    ("US-01", "Identity", "Register with Thai national ID + password",
     "Form validates 13-digit ID + checksum; creates Firebase Auth user with synthetic email; users doc auto-provisioned with role=user.",
     "Belle", "Must", 1,
     [5, 3, 5, 5, 8], [5, 5, 5, 5, 5], [5, 5, 5, 5, 5]),
    ("US-02", "Identity", "Login + logout",
     "User logs in with ID + password; logout removes FCM token from users doc.",
     "Belle", "Must", 1,
     [2, 3, 2, 2, 3], [2, 2, 2, 2, 3], [2, 2, 2, 2, 2]),
    ("US-03", "Identity", "Profile + Settings",
     "User edits name, phone, delivery pin, language preference; data persisted to users doc.",
     "Belle", "Should", 2,
     [3, 5, 3, 3, 5], [3, 3, 3, 5, 3], [3, 3, 3, 3, 3]),
    ("US-04", "Identity", "Thai ID validator + unit tests",
     "Pure-Dart + TS implementations sharing tests; rejects bad checksums.",
     "Belle", "Must", 1,
     [2, 3, 2, 3, 2], [2, 2, 2, 3, 2], [2, 2, 2, 2, 2]),
    # ── Catalog ──
    ("US-05", "Catalog", "Catalog browse + cart",
     "User filters active+in-stock items, adds to cart with quantity, sees running weight total.",
     "Bew", "Must", 1,
     [3, 5, 3, 5, 3], [3, 3, 5, 3, 3], [3, 3, 3, 3, 3]),
    ("US-06", "Catalog", "Delivery pin picker on map",
     "User drops/edits a GPS pin; defaults to profile address.",
     "Bew", "Must", 1,
     [3, 2, 3, 5, 3], [3, 3, 3, 3, 3], [3, 3, 3, 3, 3]),
    # ── Request lifecycle ──
    ("US-07", "Request", "Submit request",
     "Calls submitRequest callable; client+server validate weight ≤ max payload and stock available.",
     "Bew", "Must", 1,
     [3, 5, 3, 3, 5], [3, 3, 3, 5, 3], [3, 3, 3, 3, 3]),
    ("US-08", "Request", "Queue page (live status)",
     "Streams requests where userId == me; shows status chip + ETA when in flight.",
     "Bew", "Must", 1,
     [3, 3, 3, 5, 3], [3, 3, 3, 3, 3], [3, 3, 3, 3, 3]),
    ("US-09", "Request", "History page",
     "Lists completed / failed / cancelled requests, newest first, with timestamp.",
     "Bew", "Should", 2,
     [2, 2, 2, 3, 2], [2, 2, 2, 2, 2], [2, 2, 2, 2, 2]),
    ("US-10", "Request", "Admin Requests list with filters",
     "Admin sees all requests; filters by status + priority; live updates.",
     "Bew", "Must", 1,
     [3, 3, 3, 5, 3], [3, 3, 3, 3, 3], [3, 3, 3, 3, 3]),
    ("US-11", "Request", "Admin Request Manage: approve/reject + drone picker",
     "Admin views user profile + items; approves (stock decrement) or rejects with reason; drone picker filtered by payload + range.",
     "Bew", "Must", 1,
     [8, 5, 5, 5, 8], [5, 5, 5, 8, 5], [5, 5, 5, 5, 5]),
    ("US-12", "Catalog", "Admin Inventory + restock",
     "Admin sees per-item stock; restocks via callable; deactivates items.",
     "Bew", "Should", 2,
     [3, 3, 3, 5, 3], [3, 3, 3, 3, 3], [3, 3, 3, 3, 3]),
    # ── Fleet ──
    ("US-13", "Fleet", "Admin Drone list",
     "All drones incl. offline + maintenance; status chip + battery bar.",
     "Tawan", "Must", 1,
     [3, 2, 3, 3, 2], [3, 3, 3, 3, 3], [3, 3, 3, 3, 3]),
    ("US-14", "Fleet", "Admin Drone detail + future queue",
     "Drone status, battery, current flight, planned future flights, maintenance toggle.",
     "Tawan", "Must", 1,
     [5, 5, 8, 5, 5], [5, 5, 5, 5, 5], [5, 5, 5, 5, 5]),
    ("US-15", "Fleet", "Admin Weather panel",
     "Admin sets global weather to clear/wind/storm via callable; setting drives sim.",
     "Tawan", "Must", 1,
     [2, 3, 2, 2, 3], [2, 2, 2, 2, 3], [2, 2, 2, 2, 2]),
    # ── Tracking + maps ──
    ("US-16", "Tracking", "Tracking page with live map + interpolation",
     "flutter_map renders origin/dest/drone marker; position computed client-side every frame from flight plan.",
     "Poom", "Must", 1,
     [8, 13, 8, 8, 13], [8, 8, 8, 13, 8], [8, 8, 8, 8, 8]),
    ("US-17", "Tracking", "Confirm page + confirmDelivery wiring",
     "Banner shown when status=delivered; tap calls callable; UI updates immediately.",
     "Poom", "Must", 1,
     [2, 3, 2, 3, 2], [2, 2, 2, 2, 2], [2, 2, 2, 2, 2]),
    ("US-18", "Tracking", "Notifications inbox + FCM device registration",
     "FCM token registered to users doc; inbox shows historic events from subcollection; tap-through navigates.",
     "Poom", "Must", 1,
     [5, 5, 5, 8, 5], [5, 5, 5, 5, 5], [5, 5, 5, 5, 5]),
    ("US-19", "Tracking", "Admin Control map (all active drones)",
     "Map of every flying drone; tap a marker to see flight + request info.",
     "Poom", "Must", 1,
     [5, 5, 5, 8, 5], [5, 5, 5, 5, 5], [5, 5, 5, 5, 5]),
    # ── Backend infrastructure ──
    ("US-20", "Backend", "Firebase project setup + emulator suite",
     "Project created, Flutter + Functions wired, emulator runs Auth/Firestore/Functions locally.",
     "Aok", "Must", 1,
     [3, 5, 3, 3, 5], [3, 3, 3, 3, 3], [3, 3, 3, 3, 3]),
    ("US-21", "Backend", "Firestore rules + indexes + rules unit tests",
     "All collections have rules per spec §9; rules unit tests cover owner/admin/deny paths.",
     "Aok", "Must", 1,
     [5, 8, 5, 5, 5], [5, 5, 5, 8, 5], [5, 5, 5, 5, 5]),
    ("US-22", "Backend", "onUserCreated trigger",
     "Auth trigger writes users/{uid} with role=user, blank fcmTokens, defaults.",
     "Aok", "Must", 1,
     [1, 2, 1, 2, 1], [1, 1, 1, 2, 1], [1, 1, 1, 1, 1]),
    ("US-23", "Backend", "submitRequest callable",
     "Validates payload, weight, stock; creates requests doc; returns reqId; FCM-notifies admins.",
     "Aok", "Must", 1,
     [3, 3, 3, 5, 3], [3, 3, 3, 3, 3], [3, 3, 3, 3, 3]),
    ("US-24", "Backend", "approveRequest + rejectRequest callables",
     "approveRequest does tx: stock decrement + status change + return eligible drones. rejectRequest stores reason.",
     "Aok", "Must", 1,
     [5, 5, 5, 3, 5], [5, 5, 5, 5, 5], [5, 5, 5, 5, 5]),
    ("US-25", "Backend", "assignDrone + cancelRequest callables",
     "assignDrone creates flight, mutates drone + request statuses in tx. cancelRequest only when pending.",
     "Aok", "Must", 1,
     [5, 3, 5, 5, 5], [5, 5, 5, 5, 5], [5, 5, 5, 5, 5]),
    ("US-26", "Backend", "tickFlights v0 (movement + transitions)",
     "Scheduled every 60s; advances enroute→delivering→completed→returning→idle.",
     "Aok", "Must", 1,
     [8, 8, 5, 8, 8], [8, 8, 8, 8, 8], [8, 8, 8, 8, 8]),
    ("US-27", "Backend", "tickFlights failure dice (weather, battery, mech)",
     "Adds storm 20%, battery <15%, mechanical 1% per tick; transitions to aborted/failed with type.",
     "Aok", "Must", 1,
     [5, 5, 5, 5, 8], [5, 5, 5, 5, 5], [5, 5, 5, 5, 5]),
    ("US-28", "Backend", "confirmDelivery callable",
     "Marks request confirmed; transitions drone to returning; FCM notification.",
     "Aok", "Must", 1,
     [2, 3, 2, 2, 3], [2, 2, 2, 2, 2], [2, 2, 2, 2, 2]),
    ("US-29", "Backend", "setWeather + restockItem + toggleDroneMaintenance callables",
     "Three admin-only callables for shared state changes.",
     "Aok", "Must", 1,
     [3, 2, 3, 3, 2], [3, 3, 3, 3, 3], [3, 3, 3, 3, 3]),
    ("US-30", "Backend", "onFlightWritten trigger + FCM fan-out",
     "Firestore trigger detects state transitions; sends FCM to user + admins as appropriate.",
     "Aok", "Must", 1,
     [3, 5, 3, 5, 3], [3, 3, 3, 5, 3], [3, 3, 3, 3, 3]),
    ("US-31", "Backend", "Seed scripts (catalog, drones, admins, demo users)",
     "Idempotent npm scripts populate Firestore for demo.",
     "Aok", "Must", 1,
     [2, 3, 2, 3, 2], [2, 2, 2, 2, 2], [2, 2, 2, 2, 2]),
    # ── DevOps ──
    ("US-32", "DevOps", "GitHub Actions CI (flutter + functions + rules + gitleaks + log-presence)",
     "ci.yml runs on every PR + push to main; all gates green required to merge.",
     "Aok", "Must", 1,
     [5, 5, 5, 8, 5], [5, 5, 5, 5, 5], [5, 5, 5, 5, 5]),
    ("US-33", "DevOps", "Deploy workflows (functions + rules)",
     "On push to main, deploys functions + rules using FIREBASE_SERVICE_ACCOUNT secret.",
     "Aok", "Must", 1,
     [2, 3, 2, 2, 3], [2, 2, 2, 2, 2], [2, 2, 2, 2, 2]),
    ("US-34", "Audit", "SessionEnd hook + copy + redact scripts",
     "Each dev's session JSONL copied to docs/agent-logs/<handle>/ on session end, redacted of secrets.",
     "Belle", "Must", 1,
     [3, 5, 3, 3, 3], [3, 3, 3, 3, 3], [3, 3, 3, 3, 3]),
    ("US-35", "Audit", "Log-presence + build-log-index workflows",
     "log-presence fails PRs missing same-day log; build-log-index regenerates _index.md.",
     "Aok", "Should", 2,
     [3, 2, 3, 3, 3], [3, 3, 3, 3, 3], [3, 3, 3, 3, 3]),
    # ── Quality + demo ──
    ("US-36", "Quality", "Integration day + bug bash + demo prep",
     "Full happy-path + failure-path test on real Firebase project; seed demo data; record screencast.",
     "All",  "Must", 2,
     [5, 8, 5, 5, 8], [5, 5, 5, 8, 5], [5, 5, 5, 5, 5]),
]

ESTIMATORS = ["Aok", "Belle", "Bew", "Poom", "Tawan"]


def main():
    out = Path(__file__).resolve().parent.parent / "docs" / "02-backlog-delphi.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: Method ──
    s = wb.active
    s.title = "Method"
    method_lines = [
        ("DroneAid — Wideband Delphi Backlog Estimation", True),
        ("", False),
        ("Estimation scale: modified Fibonacci story points (1, 2, 3, 5, 8, 13).", False),
        ("One story point ≈ 0.25 person-days of focused work.", False),
        ("", False),
        ("Process applied:", True),
        ("1. Backlog items derived from the design spec acceptance criteria.", False),
        ("2. Five estimators (Aok, Belle, Bew, Poom, Tawan) independently estimate each item without seeing others' values.", False),
        ("3. Round 1: open submissions. Coordinator (Aok) computes spread (min / median / max / range).", False),
        ("4. Items with range > 2 Fibonacci steps go to discussion.", False),
        ("5. Round 2: estimators re-submit after hearing rationale for outliers.", False),
        ("6. Round 3: final convergence pass.", False),
        ("7. Consensus = median of Round 3 values, rounded to nearest Fibonacci step.", False),
        ("", False),
        ("Convergence is visible in the Backlog sheet: spread column shrinks across R1 → R2 → R3.", False),
        ("Story points roll up by epic (Summary sheet) and by owner for capacity planning.", False),
        ("", False),
        ("Estimator key:", True),
        ("E1 = Aok (Lead, Backend)", False),
        ("E2 = Belle (Identity + Shared UI)", False),
        ("E3 = Bew (Request Domain)", False),
        ("E4 = Poom (Tracking + Maps)", False),
        ("E5 = Tawan (Fleet Domain)", False),
    ]
    for r, (text, bold) in enumerate(method_lines, start=1):
        cell = s.cell(row=r, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=12)
    s.column_dimensions['A'].width = 110

    # ── Sheet 2: Backlog ──
    b = wb.create_sheet("Backlog")

    header_groups = [
        ("ID", 1),
        ("Epic", 1),
        ("User Story / Task", 1),
        ("Acceptance Criteria", 1),
        ("Owner", 1),
        ("Priority", 1),
        ("Sprint", 1),
        ("Round 1 — E1..E5", 5),
        ("R1 Min", 1), ("R1 Median", 1), ("R1 Max", 1), ("R1 Range", 1),
        ("Round 2 — E1..E5", 5),
        ("R2 Min", 1), ("R2 Median", 1), ("R2 Max", 1), ("R2 Range", 1),
        ("Round 3 — E1..E5", 5),
        ("R3 Min", 1), ("R3 Median", 1), ("R3 Max", 1), ("R3 Range", 1),
        ("Consensus (SP)", 1),
        ("Notes", 1),
    ]

    # First header row (groups)
    col = 1
    for label, span in header_groups:
        b.cell(row=1, column=col, value=label).font = Font(bold=True, color="FFFFFF")
        b.cell(row=1, column=col).fill = PatternFill("solid", start_color="1F4E78")
        b.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
        if span > 1:
            b.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
        col += span

    # Second header row (per-estimator names + min/median/max/range duplicates)
    sub_headers = (
        ["", "", "", "", "", "", ""]
        + ESTIMATORS + ["Min", "Median", "Max", "Range"]
        + ESTIMATORS + ["Min", "Median", "Max", "Range"]
        + ESTIMATORS + ["Min", "Median", "Max", "Range"]
        + ["", ""]
    )
    for i, h in enumerate(sub_headers, start=1):
        c = b.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, italic=True)
        c.alignment = Alignment(horizontal="center")
        c.fill = PatternFill("solid", start_color="D9E1F2")

    # Body
    row = 3
    for s_id, epic, story, accept, owner, pri, sprint, r1, r2, r3 in STORIES:
        b.cell(row=row, column=1, value=s_id)
        b.cell(row=row, column=2, value=epic)
        b.cell(row=row, column=3, value=story)
        b.cell(row=row, column=4, value=accept)
        b.cell(row=row, column=5, value=owner)
        b.cell(row=row, column=6, value=pri)
        b.cell(row=row, column=7, value=sprint)

        # R1 estimates cols 8..12
        for i, v in enumerate(r1):
            b.cell(row=row, column=8 + i, value=v)
        b.cell(row=row, column=13, value=f"=MIN(H{row}:L{row})")
        b.cell(row=row, column=14, value=f"=MEDIAN(H{row}:L{row})")
        b.cell(row=row, column=15, value=f"=MAX(H{row}:L{row})")
        b.cell(row=row, column=16, value=f"=O{row}-M{row}")

        # R2 estimates cols 17..21
        for i, v in enumerate(r2):
            b.cell(row=row, column=17 + i, value=v)
        b.cell(row=row, column=22, value=f"=MIN(Q{row}:U{row})")
        b.cell(row=row, column=23, value=f"=MEDIAN(Q{row}:U{row})")
        b.cell(row=row, column=24, value=f"=MAX(Q{row}:U{row})")
        b.cell(row=row, column=25, value=f"=X{row}-V{row}")

        # R3 estimates cols 26..30
        for i, v in enumerate(r3):
            b.cell(row=row, column=26 + i, value=v)
        b.cell(row=row, column=31, value=f"=MIN(Z{row}:AD{row})")
        b.cell(row=row, column=32, value=f"=MEDIAN(Z{row}:AD{row})")
        b.cell(row=row, column=33, value=f"=MAX(Z{row}:AD{row})")
        b.cell(row=row, column=34, value=f"=AG{row}-AE{row}")

        # Consensus = MEDIAN of R3 (Delphi convention)
        b.cell(row=row, column=35, value=f"=AF{row}")

        b.cell(row=row, column=36, value="")  # Notes blank

        row += 1

    # Totals row
    total_row = row
    b.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    b.cell(row=total_row, column=35, value=f"=SUM(AI3:AI{row - 1})").font = Font(bold=True)
    b.cell(row=total_row, column=35).fill = PatternFill("solid", start_color="FFF2CC")

    # Column widths
    widths = {
        1: 10, 2: 14, 3: 50, 4: 60, 5: 10, 6: 10, 7: 8,
    }
    for c, w in widths.items():
        b.column_dimensions[get_column_letter(c)].width = w
    for c in range(8, 37):
        b.column_dimensions[get_column_letter(c)].width = 8
    b.column_dimensions[get_column_letter(36)].width = 30

    # Wrap text on story + acceptance
    for r in range(3, total_row):
        b.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        b.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")

    # Conditional color scale on Range columns (green→red) to show convergence
    green_red = ColorScaleRule(
        start_type="num", start_value=0, start_color="63BE7B",
        mid_type="num", mid_value=5, mid_color="FFEB84",
        end_type="num", end_value=12, end_color="F8696B",
    )
    last_row = total_row - 1
    for col_letter in ("P", "Y", "AH"):  # R1 Range, R2 Range, R3 Range
        b.conditional_formatting.add(f"{col_letter}3:{col_letter}{last_row}", green_red)

    b.freeze_panes = "H3"

    # ── Sheet 3: Summary by Epic ──
    sm = wb.create_sheet("Summary by Epic")
    sm.cell(row=1, column=1, value="Epic").font = Font(bold=True)
    sm.cell(row=1, column=2, value="Items").font = Font(bold=True)
    sm.cell(row=1, column=3, value="Consensus SP total").font = Font(bold=True)
    sm.cell(row=1, column=4, value="Person-days est.").font = Font(bold=True)

    epics = []
    seen = set()
    for st in STORIES:
        if st[1] not in seen:
            epics.append(st[1])
            seen.add(st[1])
    for i, epic in enumerate(epics, start=2):
        sm.cell(row=i, column=1, value=epic)
        sm.cell(row=i, column=2, value=f"=COUNTIF(Backlog!B:B,A{i})")
        sm.cell(row=i, column=3, value=f"=SUMIF(Backlog!B:B,A{i},Backlog!AI:AI)")
        sm.cell(row=i, column=4, value=f"=C{i}*0.25")
    total_i = len(epics) + 2
    sm.cell(row=total_i, column=1, value="TOTAL").font = Font(bold=True)
    sm.cell(row=total_i, column=2, value=f"=SUM(B2:B{total_i - 1})").font = Font(bold=True)
    sm.cell(row=total_i, column=3, value=f"=SUM(C2:C{total_i - 1})").font = Font(bold=True)
    sm.cell(row=total_i, column=4, value=f"=SUM(D2:D{total_i - 1})").font = Font(bold=True)
    sm.column_dimensions['A'].width = 24
    for c in (2, 3, 4):
        sm.column_dimensions[get_column_letter(c)].width = 22

    # ── Sheet 4: Summary by Owner ──
    so = wb.create_sheet("Summary by Owner")
    so.cell(row=1, column=1, value="Owner").font = Font(bold=True)
    so.cell(row=1, column=2, value="Items").font = Font(bold=True)
    so.cell(row=1, column=3, value="Consensus SP").font = Font(bold=True)
    so.cell(row=1, column=4, value="Person-days").font = Font(bold=True)
    so.cell(row=1, column=5, value="Capacity (14d)").font = Font(bold=True)
    so.cell(row=1, column=6, value="Utilization").font = Font(bold=True)

    owners = ["Aok", "Belle", "Bew", "Poom", "Tawan", "All"]
    for i, ow in enumerate(owners, start=2):
        so.cell(row=i, column=1, value=ow)
        so.cell(row=i, column=2, value=f"=COUNTIF(Backlog!E:E,A{i})")
        so.cell(row=i, column=3, value=f"=SUMIF(Backlog!E:E,A{i},Backlog!AI:AI)")
        so.cell(row=i, column=4, value=f"=C{i}*0.25")
        so.cell(row=i, column=5, value=14 if ow != "All" else 0)
        so.cell(row=i, column=6, value=f"=IF(E{i}=0,\"-\",D{i}/E{i})")
        so.cell(row=i, column=6).number_format = "0.0%"

    so.column_dimensions['A'].width = 14
    for c in (2, 3, 4, 5, 6):
        so.column_dimensions[get_column_letter(c)].width = 18

    # Save
    wb.save(out)
    print(f"Wrote: {out}")


if __name__ == "__main__":
    main()
