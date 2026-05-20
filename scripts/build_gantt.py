"""Build docs/05-gantt.xlsx — DroneAid project GANTT chart, 2026-05-22 → 2026-06-04."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
from pathlib import Path


START = date(2026, 5, 22)  # Day 1
DAYS = 14                   # 2026-05-22 → 2026-06-04 inclusive

OWNER_COLORS = {
    "Aok":   "1F4E78",  # navy
    "Belle": "2E7D32",  # green
    "Bew":   "EF6C00",  # orange
    "Poom":  "6A1B9A",  # purple
    "Tawan": "00695C",  # teal
    "All":   "455A64",  # blue-grey
}
OWNER_TEXT_COLOR = "FFFFFF"

# (id, task, owner, start_day, end_day, sprint, depends_on)
# day index is 1-based against START
TASKS = [
    # Sprint 0 — bootstrap (Days 1-2)
    ("T-01", "Firebase project setup + emulator suite",                     "Aok",   1, 2,  1, ""),
    ("T-02", "Flutter scaffold: theme, go_router, Riverpod, env",           "Belle", 1, 2,  1, ""),
    ("T-03", "Data model classes + repository stubs (all features)",        "Aok",   2, 3,  1, "T-01"),
    ("T-04", "Firestore rules v0 + indexes",                                "Aok",   2, 3,  1, "T-01"),
    ("T-05", "Thai national-ID validator + unit tests",                     "Belle", 2, 3,  1, ""),

    # Sprint 1 — auth + catalog + submit
    ("T-06", "Login + register pages",                                      "Belle", 3, 4,  1, "T-05, T-02"),
    ("T-07", "onUserCreated trigger",                                       "Aok",   3, 3,  1, "T-01, T-04"),
    ("T-08", "Profile + Settings pages",                                    "Belle", 4, 5,  1, "T-06"),
    ("T-09", "Shared widgets: DroneMap, BatteryBar, StatusChip, ItemPicker", "Belle", 5, 7,  1, "T-02"),
    ("T-10", "Catalog browse + cart (user)",                                "Bew",   4, 5,  1, "T-03"),
    ("T-11", "Pin picker on map",                                           "Bew",   5, 6,  1, "T-03, T-09"),
    ("T-12", "submitRequest callable + tests",                              "Aok",   4, 5,  1, "T-03, T-04"),
    ("T-13", "Submit + Queue page (live stream)",                           "Bew",   6, 7,  1, "T-10, T-11, T-12"),
    ("T-14", "Admin Requests list (read)",                                  "Bew",   6, 7,  1, "T-03"),

    # Sprint 2 — admin manage + drone fleet + tracking
    ("T-15", "approveRequest + rejectRequest callables (stock tx)",         "Aok",   7, 8,  2, "T-12"),
    ("T-16", "Admin Request Manage page: approve/reject + drone picker",    "Bew",   7, 9,  2, "T-14, T-15"),
    ("T-17", "Admin Drone list page",                                       "Tawan", 6, 7,  1, "T-03"),
    ("T-18", "Admin Drone detail page + future queue",                      "Tawan", 7, 9,  2, "T-17"),
    ("T-19", "Drone fleet seed data",                                       "Tawan", 8, 8,  2, "T-04"),
    ("T-20", "assignDrone callable + flight document",                      "Aok",   8, 9,  2, "T-15, T-19"),
    ("T-21", "Tracking page: flutter_map + interpolation math",             "Poom",  8, 10, 2, "T-09"),
    ("T-22", "Admin Control live map",                                      "Poom",  9, 11, 2, "T-09, T-21"),

    # Sprint 3 — simulator + failures + weather + confirm
    ("T-23", "tickFlights v0: movement + state transitions",                "Aok",  10, 11, 3, "T-20"),
    ("T-24", "Confirm page + confirmDelivery wiring",                       "Poom", 10, 11, 3, "T-21"),
    ("T-25", "confirmDelivery callable",                                    "Aok",  11, 11, 3, "T-20"),
    ("T-26", "tickFlights failures: weather + battery + mechanical",        "Aok",  11, 12, 3, "T-23"),
    ("T-27", "Admin Weather panel + setWeather callable",                   "Tawan", 9, 11, 3, "T-04"),
    ("T-28", "History page (user)",                                         "Bew",  10, 11, 3, "T-13"),

    # Sprint 4 — notifications + inventory + CI
    ("T-29", "FCM device registration + token sync",                        "Poom", 11, 12, 4, "T-06"),
    ("T-30", "Notifications inbox page",                                    "Poom", 12, 12, 4, "T-29"),
    ("T-31", "onFlightWritten trigger + FCM fan-out",                       "Aok",  12, 12, 4, "T-26, T-29"),
    ("T-32", "Admin Inventory + restock callable + page",                   "Bew",  11, 12, 4, "T-15"),
    ("T-33", "toggleDroneMaintenance + cancelRequest callables",            "Aok",  12, 12, 4, "T-20"),
    ("T-34", "GitHub Actions: CI (flutter+functions+rules+gitleaks+logs)",  "Aok",   7, 9,  2, "T-04"),
    ("T-35", "GitHub Actions: deploy workflows",                            "Aok",   9, 10, 3, "T-34"),
    ("T-36", "SessionEnd hook + copy + redact-secrets scripts",             "Belle",  7, 8,  2, ""),
    ("T-37", "log-presence + build-log-index workflows",                    "Aok",  10, 11, 3, "T-34, T-36"),

    # Sprint 5 — integration + demo
    ("T-38", "Integration testing on real Firebase project (full path)",   "All",  13, 13, 5, "T-31, T-32, T-33"),
    ("T-39", "Bug bash + polish",                                          "All",  13, 14, 5, "T-38"),
    ("T-40", "Demo seed dataset + screenshots + screencast",               "All",  14, 14, 5, "T-39"),
    ("T-41", "README + presentation slides",                               "All",  14, 14, 5, "T-39"),
]


def build():
    out = Path(__file__).resolve().parent.parent / "docs" / "05-gantt.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ── Sheet: Legend ──
    leg = wb.active
    leg.title = "Legend"
    leg.cell(row=1, column=1, value="DroneAid — GANTT chart legend").font = Font(bold=True, size=13)
    leg.cell(row=2, column=1, value=f"Project window: {START} → {START + timedelta(days=DAYS - 1)} (14 days, 7-day work week)").font = Font(italic=True)
    leg.cell(row=4, column=1, value="Owner color coding").font = Font(bold=True)
    for i, (owner, color) in enumerate(OWNER_COLORS.items(), start=5):
        leg.cell(row=i, column=1, value=owner)
        cell = leg.cell(row=i, column=2, value="     ")
        cell.fill = PatternFill("solid", start_color=color)
    leg.column_dimensions['A'].width = 16
    leg.column_dimensions['B'].width = 14
    leg.cell(row=4 + len(OWNER_COLORS) + 1, column=1, value="Sprints").font = Font(bold=True)
    sprint_rows = [
        ("Sprint 0", "Bootstrap (Days 1–2)"),
        ("Sprint 1", "Auth + Catalog + Submit (Days 3–7)"),
        ("Sprint 2", "Admin manage + Fleet + Tracking core (Days 7–9)"),
        ("Sprint 3", "Simulator + Failures + Weather + Confirm (Days 9–12)"),
        ("Sprint 4", "Notifications + Inventory + CI/CD (Days 11–12)"),
        ("Sprint 5", "Integration + Demo (Days 13–14)"),
    ]
    base = 4 + len(OWNER_COLORS) + 2
    for i, (name, desc) in enumerate(sprint_rows):
        leg.cell(row=base + i, column=1, value=name).font = Font(bold=True)
        leg.cell(row=base + i, column=2, value=desc)

    # ── Sheet: GANTT ──
    g = wb.create_sheet("GANTT")

    # Header row 1: blank + day numbers
    g.cell(row=1, column=1, value="ID").font = Font(bold=True, color="FFFFFF")
    g.cell(row=1, column=2, value="Task").font = Font(bold=True, color="FFFFFF")
    g.cell(row=1, column=3, value="Owner").font = Font(bold=True, color="FFFFFF")
    g.cell(row=1, column=4, value="Sprint").font = Font(bold=True, color="FFFFFF")
    g.cell(row=1, column=5, value="Start").font = Font(bold=True, color="FFFFFF")
    g.cell(row=1, column=6, value="End").font = Font(bold=True, color="FFFFFF")
    g.cell(row=1, column=7, value="Dur").font = Font(bold=True, color="FFFFFF")
    g.cell(row=1, column=8, value="Depends on").font = Font(bold=True, color="FFFFFF")
    for col in range(1, 9):
        g.cell(row=1, column=col).fill = PatternFill("solid", start_color="1F4E78")
        g.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    DAY_COL_START = 9
    for d in range(DAYS):
        col = DAY_COL_START + d
        the_date = START + timedelta(days=d)
        c1 = g.cell(row=1, column=col, value=f"D{d + 1}")
        c1.font = Font(bold=True, color="FFFFFF")
        c1.fill = PatternFill("solid", start_color="1F4E78")
        c1.alignment = Alignment(horizontal="center")

    # Header row 2: dates + day-of-week
    g.cell(row=2, column=1, value="").fill = PatternFill("solid", start_color="D9E1F2")
    for col in range(2, 9):
        g.cell(row=2, column=col, value="").fill = PatternFill("solid", start_color="D9E1F2")
    g.cell(row=2, column=2, value="Date / DoW").font = Font(italic=True)
    for d in range(DAYS):
        the_date = START + timedelta(days=d)
        col = DAY_COL_START + d
        c2 = g.cell(row=2, column=col, value=the_date.strftime("%m-%d\n%a"))
        c2.font = Font(italic=True, size=9)
        c2.alignment = Alignment(horizontal="center", wrap_text=True)
        c2.fill = PatternFill("solid", start_color="D9E1F2")

    # Tasks rows
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    row = 3
    for t in TASKS:
        tid, name, owner, sd, ed, sprint, deps = t
        g.cell(row=row, column=1, value=tid)
        g.cell(row=row, column=2, value=name)
        g.cell(row=row, column=3, value=owner)
        g.cell(row=row, column=4, value=f"S{sprint}")
        g.cell(row=row, column=5, value=sd)
        g.cell(row=row, column=6, value=ed)
        g.cell(row=row, column=7, value=ed - sd + 1)
        g.cell(row=row, column=8, value=deps)

        # Color owner cell
        owner_color = OWNER_COLORS.get(owner, "888888")
        g.cell(row=row, column=3).fill = PatternFill("solid", start_color=owner_color)
        g.cell(row=row, column=3).font = Font(bold=True, color="FFFFFF")
        g.cell(row=row, column=3).alignment = Alignment(horizontal="center")

        # Color bar across day cells
        for d in range(sd, ed + 1):
            col = DAY_COL_START + (d - 1)
            cell = g.cell(row=row, column=col, value=tid)
            cell.fill = PatternFill("solid", start_color=owner_color)
            cell.font = Font(color="FFFFFF", size=8, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for c in range(1, DAY_COL_START + DAYS):
            g.cell(row=row, column=c).border = border

        row += 1

    # Milestone markers (demo milestones)
    milestone_row = row + 1
    g.cell(row=milestone_row, column=1, value="Milestones").font = Font(bold=True)
    milestones = [
        (5, "M1: Auth working"),
        (7, "M2: Submit + admin manage skeleton"),
        (9, "M3: Drone assignment + tracking"),
        (11, "M4: Simulator + failures live"),
        (12, "M5: Notifications + inventory"),
        (14, "M6: Demo ready"),
    ]
    for day, label in milestones:
        col = DAY_COL_START + (day - 1)
        cell = g.cell(row=milestone_row, column=col, value="◆")
        cell.font = Font(bold=True, size=14, color="C00000")
        cell.alignment = Alignment(horizontal="center")
        # Label row below
        lbl = g.cell(row=milestone_row + 1, column=col, value=label)
        lbl.font = Font(size=8, color="C00000")
        lbl.alignment = Alignment(horizontal="center", wrap_text=True)

    # Column widths
    g.column_dimensions['A'].width = 7
    g.column_dimensions['B'].width = 55
    g.column_dimensions['C'].width = 9
    g.column_dimensions['D'].width = 7
    g.column_dimensions['E'].width = 6
    g.column_dimensions['F'].width = 6
    g.column_dimensions['G'].width = 5
    g.column_dimensions['H'].width = 28
    for d in range(DAYS):
        g.column_dimensions[get_column_letter(DAY_COL_START + d)].width = 6

    g.freeze_panes = "I3"
    g.row_dimensions[2].height = 30

    # ── Sheet: Daily swimlane ──
    sw = wb.create_sheet("Daily swimlane")
    sw.cell(row=1, column=1, value="Owner / Day").font = Font(bold=True, color="FFFFFF")
    sw.cell(row=1, column=1).fill = PatternFill("solid", start_color="1F4E78")
    for d in range(DAYS):
        the_date = START + timedelta(days=d)
        col = 2 + d
        c = sw.cell(row=1, column=col, value=f"D{d + 1}\n{the_date.strftime('%m-%d %a')}")
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", start_color="1F4E78")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    owners_order = ["Aok", "Belle", "Bew", "Poom", "Tawan", "All"]
    for i, owner in enumerate(owners_order, start=2):
        sw.cell(row=i, column=1, value=owner)
        sw.cell(row=i, column=1).font = Font(bold=True, color="FFFFFF")
        sw.cell(row=i, column=1).fill = PatternFill("solid", start_color=OWNER_COLORS[owner])
        sw.cell(row=i, column=1).alignment = Alignment(horizontal="center")
        for d in range(DAYS):
            # Find tasks for this owner on this day (day index 1-based)
            day_idx = d + 1
            active = [t for t in TASKS if t[2] == owner and t[3] <= day_idx <= t[4]]
            cell = sw.cell(row=i, column=2 + d)
            if active:
                cell.value = "\n".join(t[0] for t in active)
                cell.fill = PatternFill("solid", start_color=OWNER_COLORS[owner])
                cell.font = Font(color="FFFFFF", size=8, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sw.column_dimensions['A'].width = 12
    for d in range(DAYS):
        sw.column_dimensions[get_column_letter(2 + d)].width = 10
    sw.row_dimensions[1].height = 36
    for i in range(2, 2 + len(owners_order)):
        sw.row_dimensions[i].height = 50
    sw.freeze_panes = "B2"

    # ── Sheet: Per-owner workload count ──
    wl = wb.create_sheet("Workload")
    wl.cell(row=1, column=1, value="Owner").font = Font(bold=True, color="FFFFFF")
    wl.cell(row=1, column=2, value="Tasks").font = Font(bold=True, color="FFFFFF")
    wl.cell(row=1, column=3, value="Person-days").font = Font(bold=True, color="FFFFFF")
    wl.cell(row=1, column=4, value="Capacity (14d)").font = Font(bold=True, color="FFFFFF")
    wl.cell(row=1, column=5, value="Utilization").font = Font(bold=True, color="FFFFFF")
    for c in range(1, 6):
        wl.cell(row=1, column=c).fill = PatternFill("solid", start_color="1F4E78")
        wl.cell(row=1, column=c).alignment = Alignment(horizontal="center")

    for i, owner in enumerate(owners_order, start=2):
        wl.cell(row=i, column=1, value=owner)
        owner_tasks = [t for t in TASKS if t[2] == owner]
        wl.cell(row=i, column=2, value=len(owner_tasks))
        days_used = sum(t[4] - t[3] + 1 for t in owner_tasks)
        wl.cell(row=i, column=3, value=days_used)
        wl.cell(row=i, column=4, value=14 if owner != "All" else 0)
        wl.cell(row=i, column=5, value=f"=IF(D{i}=0,\"-\",C{i}/D{i})")
        wl.cell(row=i, column=5).number_format = "0.0%"
        wl.cell(row=i, column=1).fill = PatternFill("solid", start_color=OWNER_COLORS[owner])
        wl.cell(row=i, column=1).font = Font(bold=True, color="FFFFFF")
        wl.cell(row=i, column=1).alignment = Alignment(horizontal="center")

    for col, w in {1: 12, 2: 10, 3: 14, 4: 16, 5: 14}.items():
        wl.column_dimensions[get_column_letter(col)].width = w

    wb.save(out)
    print(f"Wrote: {out}")


if __name__ == "__main__":
    build()
